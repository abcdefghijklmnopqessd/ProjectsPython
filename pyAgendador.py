import streamlit as st 
import openpyxl 
import time 
import schedule # Modulo de agendamento
import threading
import datetime
import pytz

dias_semana = {
    "Segunda-feira": "monday",
    "Terça-feira": "tuesday",
    "Quarta-feira": "wednesday",
    "Quinta-feira": "thursday",
    "Sexta-feira": "friday"
}


def criar_arquivo_excel():
  try:
    wb = openpyxl.load_workbook("Agendamentos.xlsx")
  except FileNotFoundError:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Agendamentos"
    ws.append(["Dia", "Horário"])
    wb.save("Agendamentos.xlsx")


def salvar_agendamentos_excel(dia, horario):
  wb = openpyxl.load_workbook("Agendamentos.xlsx")
  ws = wb.active
  ws.append([dia,horario])
  wb.save("Agendamentos.xlsx")

def carregar_agendamentos_excel():
    import re
    wb = openpyxl.load_workbook("Agendamentos.xlsx")
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        dia, horario = row
        # Converter horário para formato correto se for apenas um número
        if isinstance(horario, (int, float)):
            horario = f"{int(horario):02d}:00"
        
        # Garantir que o horário seja string
        horario = str(horario)
        
        # Se o horário contém apenas horas, adicionar minutos
        if re.match(r'^([01]?[0-9]|2[0-3])$', horario):
            horario = f"{int(horario):02d}:00"
        
        # Verificar se é um dia válido e um formato de hora válido
        if dia in ["monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday"] and re.match(r'^([01]?\d|2[0-3]):([0-5]\d)$', horario):
            # Garantir formato correto HH:MM com zeros à esquerda
            hora, minuto = horario.split(':')
            horario_formatado = f"{int(hora):02d}:{int(minuto):02d}"
            
            try:
                schedule.every().__getattribute__(dia).at(horario_formatado).do(tarefa)
                print(f"Agendamento carregado: {dia} às {horario_formatado}")
            except Exception as e:
                print(f"Erro ao carregar agendamento {dia} às {horario_formatado}: {str(e)}")
        else:
            print(f"Formato inválido no agendamento: {dia} às {horario}")

def tarefa():
    # Obter hora atual em UTC
    hora_utc = datetime.datetime.now(pytz.UTC)
    # Converter para o fuso horário de Brasília
    fuso_brasil = pytz.timezone('America/Sao_Paulo')
    hora_brasil = hora_utc.astimezone(fuso_brasil)
    
    print(f"Executando tarefa! Hora UTC: {hora_utc.strftime('%H:%M')}, Hora Brasil: {hora_brasil.strftime('%H:%M')}")
    
    # Mostrar vibração com JavaScript no Streamlit
    js_code = """
    <script>
        if (navigator.vibrate) {
            navigator.vibrate(1000);  // Vibra por 1 segundo
        } else {
            alert('Vibração não suportada no navegador.');
        }
    </script>
    """
    st.components.v1.html(js_code)
    st.success(f"Tarefa executada com sucesso! Hora local do servidor: {hora_brasil.strftime('%H:%M')}")

def run_scheduler():
  while True:
    schedule.run_pending()
    time.sleep(1)


criar_arquivo_excel()

carregar_agendamentos_excel()

threading.Thread(target=run_scheduler, daemon=True).start()


st.title("📅 Agendador de Tarefas")

dia_escolhido = st.selectbox("Escolha o dia da semana:", list(dias_semana.keys()))
horario = st.text_input("Digite o horario (HH:MM)")

if st.button("Agendar"):
  # Validar e formatar o horário
  import re
  
  # Se o usuário inseriu apenas um número (como "9"), converter para "09:00"
  if re.match(r'^([01]?[0-9]|2[0-3])$', horario):
    horario = f"{int(horario):02d}:00"
    
  # Verificar formato HH:MM (mais flexível, aceita 9:30 além de 09:30)
  if re.match(r'^([01]?[0-9]|2[0-3]):([0-5][0-9])$', horario):
    # Garantir formato padronizado com zeros à esquerda
    hora, minuto = horario.split(':')
    horario_formatado = f"{int(hora):02d}:{int(minuto):02d}"
    
    dia_em_ingles = dias_semana[dia_escolhido]
    salvar_agendamentos_excel(dia_em_ingles, horario_formatado)
    # Agendar a tarefa imediatamente
    schedule.every().__getattribute__(dia_em_ingles).at(horario_formatado).do(tarefa)
    st.success(f"Tarefa agendada para {dia_escolhido} às {horario_formatado}")
  else:
    st.error("Formato de hora inválido! Use o formato HH:MM (exemplo: 14:30) ou apenas a hora (exemplo: 9)")


def excluir_agendamento(index):
    # Carregar o arquivo Excel
    wb = openpyxl.load_workbook("Agendamentos.xlsx")
    ws = wb.active
    
    # Os índices começam do 0, mas as linhas do Excel começam do 1,
    # e a primeira linha é o cabeçalho, então adicionamos 2
    linha_para_excluir = index + 2
    
    # Excluir a linha
    ws.delete_rows(linha_para_excluir, 1)
    
    # Salvar as alterações
    wb.save("Agendamentos.xlsx")
    
    # Recarregar os agendamentos
    schedule.clear()
    carregar_agendamentos_excel()
    
    # Forçar atualização da interface
    st.rerun()

st.subheader("📌 Agendamentos Atuais")
wb = openpyxl.load_workbook("Agendamentos.xlsx")
ws = wb.active

# Armazenar todos os agendamentos para mostrar na interface
agendamentos = []
for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
    dia, horario = row
    try:
        dia_em_portugues = [key for key, value in dias_semana.items() if value == dia][0]
        agendamentos.append((i, dia_em_portugues, horario))
    except IndexError:
        # Caso o dia não esteja no dicionário, mostrar o valor original
        agendamentos.append((i, dia, horario))

# Se não houver agendamentos, mostrar uma mensagem
if not agendamentos:
    st.info("Nenhum agendamento cadastrado ainda.")
else:
    # Mostrar cada agendamento com um botão de exclusão
    for index, dia, horario in agendamentos:
        col1, col2 = st.columns([4, 1])
        with col1:
            st.write(f"✅ {dia} às {horario}")
        with col2:
            # Usar um nome único para cada botão
            if st.button("🗑️ Excluir", key=f"del_{index}"):
                excluir_agendamento(index)